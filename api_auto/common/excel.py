import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

class DoExcel:
    def __init__(self,file_path,sheet_name):
        self.file_path=file_path
        self.sheet_name=sheet_name

    #读取excel数据
    def read_excel_dict(self):
        """得到每一行数据是字典。"""
        wb = openpyxl.load_workbook(self.file_path)
        sheet: Worksheet = wb[self.sheet_name]
        rows = list(sheet.values)
        title = rows[0]  # ['id', 'title', 'url', ...]
        rows = rows[1:]
        new_rows = [dict(zip(title, row)) for row in rows]
        # wb.close()  #关闭excel
        return new_rows

    #写入数据到excel
    def write_res(self,row,column,test_res):   #row:excel行  column:excel列  final_res：写入结果
        wb = openpyxl.load_workbook(self.file_path)
        sheet = wb[self.sheet_name]
        sheet.cell(row,column).value = test_res
        wb.save(self.file_path)
        wb.close()


